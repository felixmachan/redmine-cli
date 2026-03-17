from __future__ import annotations

import calendar
import json
import os
import shutil
import subprocess
from collections import defaultdict
from datetime import date, datetime, timedelta
from pathlib import Path
from urllib.error import HTTPError, URLError
from urllib.parse import urlencode
from urllib.request import Request, urlopen

from openpyxl import load_workbook

from ..config import AppConfig
from ..models import TimetableRunResult
from ..workspace import resolve_workspace_path

START_ROW = 11
END_ROW = 42
SUMMARY_ROW = 43
DATE_COL = 1
ARR_COL = 2
DEP_COL = 3
HOURS_COL = 4


def list_recent_months(today: date, count: int = 18) -> list[tuple[str, int, int]]:
    months: list[tuple[str, int, int]] = []
    cursor = today.replace(day=1)
    for _ in range(count):
        months.append((cursor.strftime("%Y %B"), cursor.year, cursor.month))
        cursor = (cursor - timedelta(days=1)).replace(day=1)
    return months


def month_to_date_range(year: int, month: int) -> tuple[str, str]:
    last_day = calendar.monthrange(year, month)[1]
    return f"{year:04d}-{month:02d}-01", f"{year:04d}-{month:02d}-{last_day:02d}"


def resolve_date_range(today: date) -> tuple[str, str]:
    manual_from = os.getenv("DATE_FROM", "").strip()
    manual_to = os.getenv("DATE_TO", "").strip()
    if manual_from and manual_to:
        return manual_from, manual_to

    if today.day == 1:
        prev_month_last_day = today - timedelta(days=1)
        start = prev_month_last_day.replace(day=1)
        end = prev_month_last_day
        return start.isoformat(), end.isoformat()

    start = today.replace(day=1)
    end = today
    return start.isoformat(), end.isoformat()


def get_time_entries(base_url: str, api_key: str, user_id: str, date_from: str, date_to: str):
    def fetch_page_with_curl(params: dict):
        url = f"{base_url.rstrip('/')}/time_entries.json?{urlencode(params)}"
        cmd = [
            "curl",
            "-sS",
            "-X",
            "GET",
            url,
            "-H",
            f"X-Redmine-API-Key: {api_key}",
            "-H",
            "Content-Type: application/json",
            "-H",
            "Accept: application/json",
        ]
        result = subprocess.run(cmd, capture_output=True)
        if result.returncode != 0:
            stderr_text = (result.stderr or b"").decode("utf-8", errors="replace")
            raise RuntimeError(f"curl request failed ({result.returncode}): {stderr_text.strip()}")
        stdout_text = (result.stdout or b"").decode("utf-8", errors="replace")
        try:
            return json.loads(stdout_text)
        except Exception as exc:
            raise RuntimeError(f"curl returned non-JSON response: {stdout_text[:400]}") from exc

    def fetch_page_with_urllib(params: dict):
        url = f"{base_url.rstrip('/')}/time_entries.json?{urlencode(params)}"
        common_headers = {
            "Accept": "application/json",
            "Content-Type": "application/json",
            "User-Agent": os.getenv("HTTP_USER_AGENT", "curl/8.5.0"),
        }
        attempts = [
            Request(url, headers={**common_headers, "X-Redmine-API-Key": api_key}),
            Request(f"{url}&{urlencode({'key': api_key})}", headers=common_headers),
        ]
        payload = None
        last_error = None
        for req in attempts:
            try:
                with urlopen(req) as response:
                    payload = json.loads(response.read().decode("utf-8"))
                break
            except HTTPError as exc:
                body = ""
                try:
                    body = exc.read().decode("utf-8", errors="replace")
                except Exception:
                    body = "<no response body>"
                last_error = RuntimeError(
                    f"Redmine API HTTP {exc.code} for {req.full_url}\nResponse: {body}\n"
                    "Check REDMINE_BASE_URL / REDMINE_API_KEY / REDMINE_USER_ID."
                )
            except URLError as exc:
                last_error = RuntimeError(
                    f"Network error while calling Redmine API: {exc}\n"
                    "Check REDMINE_BASE_URL and network connectivity."
                )
        if payload is None:
            raise last_error if last_error is not None else RuntimeError("Unknown Redmine API error.")
        return payload

    def fetch_for_user_filter(user_filter: str | None):
        all_entries = []
        offset = 0
        limit = 100
        use_curl = os.getenv("USE_CURL", "1").strip().lower() not in {"0", "false", "no"}
        while True:
            params = {"from": date_from, "to": date_to, "limit": limit, "offset": offset}
            if user_filter:
                params["user_id"] = user_filter
            payload = fetch_page_with_curl(params) if use_curl else fetch_page_with_urllib(params)
            batch = payload.get("time_entries", [])
            all_entries.extend(batch)
            total_count = int(payload.get("total_count", len(all_entries)))
            offset += len(batch)
            if not batch or offset >= total_count:
                break
        return all_entries

    user_filters = []
    if user_id.strip():
        user_filters.append(user_id.strip())
    if "me" not in user_filters:
        user_filters.append("me")
    user_filters.append(None)

    for user_filter in user_filters:
        entries = fetch_for_user_filter(user_filter)
        if entries:
            return entries
    return []


def aggregate_hours_by_day(entries):
    by_day = defaultdict(float)
    for entry in entries:
        spent_on = entry.get("spent_on")
        hours = float(entry.get("hours", 0))
        if spent_on:
            by_day[spent_on] += hours
    return sorted(by_day.items(), key=lambda item: item[0])


def fill_excel(
    excel_in: str,
    excel_out: str,
    sheet_name: str,
    days,
    arrival_time_str: str,
    print_area: str,
) -> str:
    workbook = load_workbook(excel_in)
    worksheet = workbook[sheet_name] if sheet_name in workbook.sheetnames else workbook.active
    arrival_time = datetime.strptime(arrival_time_str, "%H:%M").time()

    for row in range(START_ROW, END_ROW + 1):
        worksheet.row_dimensions[row].hidden = False
        worksheet.cell(row, DATE_COL).value = None
        worksheet.cell(row, ARR_COL).value = None
        worksheet.cell(row, DEP_COL).value = None
        worksheet.cell(row, HOURS_COL).value = None

    max_rows = END_ROW - START_ROW + 1
    if len(days) > max_rows:
        raise ValueError(f"Too many workdays. Need {len(days)} rows, available {max_rows}.")

    row = START_ROW
    total_worked_minutes = 0
    for day_str, total_hours in days:
        day_date = datetime.strptime(day_str, "%Y-%m-%d").date()
        worked_minutes = int(round(total_hours * 60))
        total_worked_minutes += worked_minutes
        dep_time = (datetime.combine(day_date, arrival_time) + timedelta(minutes=worked_minutes)).time()
        hours_decimal = round(worked_minutes / 60, 2)

        worksheet.cell(row, DATE_COL).value = day_date
        worksheet.cell(row, ARR_COL).value = arrival_time
        worksheet.cell(row, DEP_COL).value = dep_time
        worksheet.cell(row, HOURS_COL).value = hours_decimal

        worksheet.cell(row, DATE_COL).number_format = "yyyy-mm-dd"
        worksheet.cell(row, ARR_COL).number_format = "h:mm"
        worksheet.cell(row, DEP_COL).number_format = "h:mm"
        worksheet.cell(row, HOURS_COL).number_format = "0.00"
        row += 1

    for row_idx in range(START_ROW + len(days), END_ROW + 1):
        worksheet.row_dimensions[row_idx].hidden = True

    total_hours_decimal = round(total_worked_minutes / 60, 2)
    worksheet.cell(SUMMARY_ROW, HOURS_COL).value = total_hours_decimal
    worksheet.cell(SUMMARY_ROW, HOURS_COL).number_format = "0.00"
    worksheet.cell(SUMMARY_ROW, 9).value = f"=D{SUMMARY_ROW}*3800"
    worksheet.print_area = print_area
    workbook.save(excel_out)
    return print_area


def export_pdf(
    excel_path: str,
    pdf_path: str,
    sheet_name: str,
    print_area: str,
    allow_libreoffice_fallback: bool,
) -> str:
    excel_error = None
    if os.name == "nt":
        try:
            import win32com.client as win32  # type: ignore

            excel = None
            workbook = None
            try:
                excel = win32.DispatchEx("Excel.Application")
                excel.Visible = False
                excel.DisplayAlerts = False
                workbook = excel.Workbooks.Open(os.path.abspath(excel_path))
                worksheet = workbook.Worksheets(sheet_name) if sheet_name else workbook.Worksheets(1)
                worksheet.PageSetup.PrintArea = print_area
                worksheet.ExportAsFixedFormat(0, os.path.abspath(pdf_path), 0, True, False)
                return "excel"
            finally:
                if workbook is not None:
                    try:
                        workbook.Close(SaveChanges=False)
                    except Exception:
                        pass
                if excel is not None:
                    excel.Quit()
        except Exception as exc:
            excel_error = exc
            if not allow_libreoffice_fallback:
                raise RuntimeError(
                    "Excel PDF export failed on Windows. Install/repair Excel+pywin32, "
                    "or enable ALLOW_LIBREOFFICE_FALLBACK_ON_WINDOWS."
                ) from exc

    office_bin = shutil.which("soffice") or shutil.which("libreoffice")
    if not office_bin:
        if excel_error is not None:
            raise RuntimeError("Excel PDF export failed and LibreOffice fallback is not available.") from excel_error
        raise RuntimeError("LibreOffice is not installed. Install it or disable PDF export.")

    excel_abs = Path(excel_path).resolve()
    pdf_abs = Path(pdf_path).resolve()
    cmd = [office_bin, "--headless", "--convert-to", "pdf", "--outdir", str(pdf_abs.parent), str(excel_abs)]
    proc = subprocess.run(cmd, capture_output=True, text=True)
    if proc.returncode != 0:
        raise RuntimeError(
            "LibreOffice PDF conversion failed.\n"
            f"Command: {' '.join(cmd)}\n"
            f"stdout: {proc.stdout}\n"
            f"stderr: {proc.stderr}"
        )

    generated = pdf_abs.parent / f"{excel_abs.stem}.pdf"
    if not generated.exists():
        raise RuntimeError("LibreOffice finished without output PDF.")
    if generated.resolve() != pdf_abs:
        generated.replace(pdf_abs)
    return "libreoffice"


def run_timetable(
    config: AppConfig,
    date_from: str,
    date_to: str,
    pdf_out: str | None = None,
    excel_out: str | None = None,
) -> TimetableRunResult:
    if not config.redmine.base_url or not config.redmine.api_key or not config.redmine.user_id:
        raise RuntimeError("Missing REDMINE_BASE_URL, REDMINE_API_KEY, or REDMINE_USER_ID.")

    excel_in_path = str(resolve_workspace_path(config.workspace_root, config.timetable.excel_in))

    excel_out_candidate = Path(excel_out or config.timetable.excel_out)
    pdf_out_candidate = Path(pdf_out or config.timetable.pdf_out)
    excel_out_path = str(excel_out_candidate if excel_out_candidate.is_absolute() else config.current_dir / excel_out_candidate)
    pdf_out_path = str(pdf_out_candidate if pdf_out_candidate.is_absolute() else config.current_dir / pdf_out_candidate)

    entries = get_time_entries(
        config.redmine.base_url,
        config.redmine.api_key,
        config.redmine.user_id,
        date_from,
        date_to,
    )
    days = aggregate_hours_by_day(entries)
    effective_print_area = fill_excel(
        excel_in_path,
        excel_out_path,
        config.timetable.sheet_name,
        days,
        config.timetable.arrival_time,
        config.timetable.print_area,
    )
    pdf_engine = export_pdf(
        excel_out_path,
        pdf_out_path,
        config.timetable.sheet_name,
        effective_print_area,
        config.timetable.allow_libreoffice_fallback_on_windows,
    )

    return TimetableRunResult(
        excel_path=excel_out_path,
        pdf_path=pdf_out_path,
        pdf_engine=pdf_engine,
        date_from=date_from,
        date_to=date_to,
        print_area=effective_print_area,
        fetched_entries=len(entries),
        filled_workdays=len(days),
    )
