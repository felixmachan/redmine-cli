import calendar
import json
import os
import shutil
import subprocess
import sys
from collections import defaultdict
from datetime import date, datetime, timedelta
from pathlib import Path
from urllib.parse import urlencode
from urllib.error import HTTPError, URLError
from urllib.request import Request, urlopen

from openpyxl import load_workbook

START_ROW = 11
END_ROW = 42
SUMMARY_ROW = 43
DATE_COL = 1  # A
ARR_COL = 2  # B
DEP_COL = 3  # C
HOURS_COL = 4  # D


def app_root() -> Path:
    if getattr(sys, "frozen", False):
        return Path(sys.executable).resolve().parent
    return Path(__file__).resolve().parent


def repo_root() -> Path:
    return app_root().parent


def resolve_path(path_value: str) -> Path:
    p = Path(path_value)
    if p.is_absolute():
        return p
    app_candidate = app_root() / p
    if app_candidate.exists():
        return app_candidate

    repo_candidate = repo_root() / p
    if repo_candidate.exists():
        return repo_candidate

    return app_candidate


def load_env_file(path: str = ".env") -> None:
    candidates = [app_root() / path]
    if not getattr(sys, "frozen", False):
        candidates.append(repo_root() / path)

    for env_path in candidates:
        if not env_path.exists():
            continue
        with open(env_path, "r", encoding="utf-8") as f:
            for raw_line in f:
                line = raw_line.strip()
                if not line or line.startswith("#") or "=" not in line:
                    continue
                key, value = line.split("=", 1)
                key = key.strip()
                value = value.strip().strip("'").strip('"')
                if key and key not in os.environ:
                    os.environ[key] = value
        return


def getenv_required(name: str) -> str:
    value = os.getenv(name, "").strip()
    if not value:
        raise ValueError(f"Missing required env var: {name}")
    return value


def resolve_date_range(today: date) -> tuple[str, str]:
    # Optional manual override for debugging.
    manual_from = os.getenv("DATE_FROM", "").strip()
    manual_to = os.getenv("DATE_TO", "").strip()
    if manual_from and manual_to:
        return manual_from, manual_to

    # Production behavior:
    # - On day 1, process full previous month.
    # - On other days, process current month up to today.
    if today.day == 1:
        prev_month_last_day = today - timedelta(days=1)
        start = prev_month_last_day.replace(day=1)
        end = prev_month_last_day
        return start.isoformat(), end.isoformat()

    start = today.replace(day=1)
    end = today
    return start.isoformat(), end.isoformat()


def choose_month_windows_tk(today: date, default_pdf_out: str) -> tuple[str, str, str] | None:
    if os.name != "nt":
        return None
    if os.getenv("WINDOWS_MONTH_PICKER", "1").strip().lower() in {"0", "false", "no"}:
        return None

    try:
        import tkinter as tk
        from tkinter import ttk
    except Exception:
        return None

    selected = {"value": None}
    root = tk.Tk()
    root.title("Redmine timetable export")
    root.resizable(False, False)

    frame = ttk.Frame(root, padding=12)
    frame.grid(row=0, column=0)

    ttk.Label(frame, text="Export honap:").grid(row=0, column=0, columnspan=2, sticky="w")

    year_var = tk.IntVar(value=today.year)
    month_var = tk.IntVar(value=today.month)

    year_box = ttk.Spinbox(frame, from_=2020, to=2100, textvariable=year_var, width=8)
    month_box = ttk.Spinbox(frame, from_=1, to=12, textvariable=month_var, width=5)
    year_box.grid(row=1, column=0, padx=(0, 8), pady=(6, 10))
    month_box.grid(row=1, column=1, pady=(6, 10))

    ttk.Label(frame, text="Kimeneti PDF nev:").grid(row=2, column=0, columnspan=2, sticky="w")
    pdf_var = tk.StringVar(value=default_pdf_out)
    pdf_entry = ttk.Entry(frame, textvariable=pdf_var, width=28)
    pdf_entry.grid(row=3, column=0, columnspan=2, sticky="we", pady=(6, 10))

    def on_ok() -> None:
        year = year_var.get()
        month = month_var.get()
        last_day = calendar.monthrange(year, month)[1]
        pdf_name = pdf_var.get().strip() or default_pdf_out
        if not pdf_name.lower().endswith(".pdf"):
            pdf_name += ".pdf"
        selected["value"] = (
            f"{year:04d}-{month:02d}-01",
            f"{year:04d}-{month:02d}-{last_day:02d}",
            pdf_name,
        )
        root.destroy()

    def on_cancel() -> None:
        root.destroy()

    btn_frame = ttk.Frame(frame)
    btn_frame.grid(row=4, column=0, columnspan=2, sticky="e")
    ttk.Button(btn_frame, text="OK", command=on_ok).grid(row=0, column=0, padx=(0, 6))
    ttk.Button(btn_frame, text="Cancel", command=on_cancel).grid(row=0, column=1)

    root.protocol("WM_DELETE_WINDOW", on_cancel)
    root.mainloop()
    return selected["value"]


def choose_pdf_save_path_windows(default_pdf_out: str) -> str:
    if os.name != "nt":
        return default_pdf_out
    if os.getenv("PDF_SAVE_DIALOG", "1").strip().lower() in {"0", "false", "no"}:
        return default_pdf_out

    try:
        import tkinter as tk
        from tkinter import filedialog
    except Exception:
        return default_pdf_out

    default_path = resolve_path(default_pdf_out)
    root = tk.Tk()
    root.withdraw()
    root.attributes("-topmost", True)
    selected = filedialog.asksaveasfilename(
        title="PDF mentes helye",
        defaultextension=".pdf",
        filetypes=[("PDF files", "*.pdf")],
        initialdir=str(default_path.parent),
        initialfile=default_path.name,
    )
    root.destroy()
    return selected.strip() if selected else default_pdf_out


def get_time_entries(base_url: str, api_key: str, user_id: str, date_from: str, date_to: str):
    def fetch_page_with_curl(params: dict):
        base_api_url = f"{base_url.rstrip('/')}/time_entries.json"
        query = urlencode(params)
        url = f"{base_api_url}?{query}"
        cmd = [
            "curl",
            "-sS",
            "-X",
            "GET",
            url,
            "-H",
            f"X-Redmine-API-Key: {api_key}",
            "-H",
            "Content-Type: application/json",
            "-H",
            "Accept: application/json",
        ]
        result = subprocess.run(cmd, capture_output=True)
        if result.returncode != 0:
            stderr_text = (result.stderr or b"").decode("utf-8", errors="replace")
            raise RuntimeError(f"curl request failed ({result.returncode}): {stderr_text.strip()}")
        stdout_bytes = result.stdout or b""
        stdout_text = stdout_bytes.decode("utf-8", errors="replace")
        try:
            return json.loads(stdout_text)
        except Exception as exc:
            raise RuntimeError(f"curl returned non-JSON response: {stdout_text[:400]}") from exc

    def fetch_page_with_urllib(params: dict):
        base_api_url = f"{base_url.rstrip('/')}/time_entries.json"
        url = f"{base_api_url}?{urlencode(params)}"
        user_agent = os.getenv("HTTP_USER_AGENT", "curl/8.5.0")
        common_headers = {
            "Accept": "application/json",
            "Content-Type": "application/json",
            "User-Agent": user_agent,
        }
        attempts = [
            Request(url, headers={**common_headers, "X-Redmine-API-Key": api_key}),
            Request(f"{url}&{urlencode({'key': api_key})}", headers=common_headers),
        ]
        payload = None
        last_error = None
        for req in attempts:
            try:
                with urlopen(req) as resp:
                    payload = json.loads(resp.read().decode("utf-8"))
                break
            except HTTPError as exc:
                body = ""
                try:
                    body = exc.read().decode("utf-8", errors="replace")
                except Exception:
                    body = "<no response body>"
                last_error = RuntimeError(
                    f"Redmine API HTTP {exc.code} for {req.full_url}\n"
                    f"Response: {body}\n"
                    "Check REDMINE_BASE_URL / REDMINE_API_KEY / REDMINE_USER_ID."
                )
            except URLError as exc:
                last_error = RuntimeError(
                    f"Network error while calling Redmine API: {exc}\n"
                    "Check REDMINE_BASE_URL and network connectivity."
                )
        if payload is None:
            raise last_error if last_error is not None else RuntimeError("Unknown Redmine API error.")
        return payload

    def fetch_for_user_filter(user_filter: str | None):
        all_entries = []
        offset = 0
        limit = 100
        use_curl = os.getenv("USE_CURL", "1").strip().lower() not in {"0", "false", "no"}

        while True:
            params = {
                "from": date_from,
                "to": date_to,
                "limit": limit,
                "offset": offset,
            }
            if user_filter:
                params["user_id"] = user_filter

            payload = fetch_page_with_curl(params) if use_curl else fetch_page_with_urllib(params)

            batch = payload.get("time_entries", [])
            all_entries.extend(batch)

            total_count = int(payload.get("total_count", len(all_entries)))
            offset += len(batch)
            if len(batch) == 0 or offset >= total_count:
                break

        return all_entries

    user_filters = []
    if user_id.strip():
        user_filters.append(user_id.strip())
    if "me" not in user_filters:
        user_filters.append("me")
    user_filters.append(None)

    for idx, user_filter in enumerate(user_filters):
        entries = fetch_for_user_filter(user_filter)
        if entries:
            print(f"Using Redmine filter user_id={user_filter if user_filter else '<none>'}")
            return entries
        if idx < len(user_filters) - 1:
            print(f"No entries with user_id={user_filter if user_filter else '<none>'}, trying fallback...")

    return []


def aggregate_hours_by_day(entries):
    by_day = defaultdict(float)
    for entry in entries:
        spent_on = entry.get("spent_on")
        hours = float(entry.get("hours", 0))
        if spent_on:
            by_day[spent_on] += hours
    return sorted(by_day.items(), key=lambda item: item[0])


def fill_excel(
    excel_in: str,
    excel_out: str,
    sheet_name: str,
    days,
    arrival_time_str: str,
    print_area: str,
) -> str:
    wb = load_workbook(excel_in)
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active
    arrival_time = datetime.strptime(arrival_time_str, "%H:%M").time()

    for row in range(START_ROW, END_ROW + 1):
        ws.row_dimensions[row].hidden = False
        ws.cell(row, DATE_COL).value = None
        ws.cell(row, ARR_COL).value = None
        ws.cell(row, DEP_COL).value = None
        ws.cell(row, HOURS_COL).value = None

    max_rows = END_ROW - START_ROW + 1
    if len(days) > max_rows:
        raise ValueError(
            f"Too many workdays for A{START_ROW}:A{END_ROW}. "
            f"Need {len(days)} rows, available {max_rows}."
        )

    row = START_ROW
    total_worked_minutes = 0
    for day_str, total_hours in days:
        day_date = datetime.strptime(day_str, "%Y-%m-%d").date()

        # Compute times from minutes to avoid float precision artifacts.
        worked_minutes = int(round(total_hours * 60))
        total_worked_minutes += worked_minutes
        dep_time = (datetime.combine(day_date, arrival_time) + timedelta(minutes=worked_minutes)).time()
        hours_decimal = round(worked_minutes / 60, 2)

        ws.cell(row, DATE_COL).value = day_date
        ws.cell(row, ARR_COL).value = arrival_time
        ws.cell(row, DEP_COL).value = dep_time
        ws.cell(row, HOURS_COL).value = hours_decimal

        ws.cell(row, DATE_COL).number_format = "yyyy-mm-dd"
        ws.cell(row, ARR_COL).number_format = "h:mm"
        ws.cell(row, DEP_COL).number_format = "h:mm"
        ws.cell(row, HOURS_COL).number_format = "0.00"

        row += 1

    # Hide unused attendance rows to keep lower static section unchanged.
    for row_idx in range(START_ROW + len(days), END_ROW + 1):
        ws.row_dimensions[row_idx].hidden = True

    # Write total directly instead of relying on Excel formula (avoids circular refs).
    total_hours_decimal = round(total_worked_minutes / 60, 2)
    ws.cell(SUMMARY_ROW, HOURS_COL).value = total_hours_decimal
    ws.cell(SUMMARY_ROW, HOURS_COL).number_format = "0.00"
    ws.cell(SUMMARY_ROW, 9).value = f"=D{SUMMARY_ROW}*3800"

    ws.print_area = print_area

    wb.save(excel_out)
    return print_area

def export_pdf(excel_path: str, pdf_path: str, sheet_name: str, print_area: str) -> str:
    excel_error = None
    if os.name == "nt":
        try:
            import win32com.client as win32  # type: ignore

            excel = None
            workbook = None
            try:
                excel = win32.DispatchEx("Excel.Application")
                excel.Visible = False
                excel.DisplayAlerts = False

                workbook = excel.Workbooks.Open(os.path.abspath(excel_path))
                worksheet = workbook.Worksheets(sheet_name) if sheet_name else workbook.Worksheets(1)
                worksheet.PageSetup.PrintArea = print_area
                worksheet.ExportAsFixedFormat(
                    0,  # xlTypePDF
                    os.path.abspath(pdf_path),
                    0,  # xlQualityStandard
                    True,
                    False,
                )
                return "excel"
            finally:
                if workbook is not None:
                    try:
                        workbook.Close(SaveChanges=False)
                    except Exception:
                        pass
                if excel is not None:
                    excel.Quit()
        except Exception as exc:
            excel_error = exc
            allow_windows_fallback = os.getenv(
                "ALLOW_LIBREOFFICE_FALLBACK_ON_WINDOWS", "0"
            ).strip().lower() in {"1", "true", "yes"}
            if not allow_windows_fallback:
                raise RuntimeError(
                    "Excel PDF export failed on Windows. "
                    "Install/repair Excel+pywin32, or set ALLOW_LIBREOFFICE_FALLBACK_ON_WINDOWS=1."
                ) from exc

    # Linux fallback: LibreOffice headless conversion.
    office_bin = shutil.which("soffice") or shutil.which("libreoffice")
    if not office_bin:
        if excel_error is not None:
            raise RuntimeError(
                "Excel PDF export failed and LibreOffice fallback is not available."
            ) from excel_error
        raise RuntimeError(
            "LibreOffice is not installed. Install it (e.g. 'apt-get install libreoffice-calc') or disable PDF export."
        )

    excel_abs = Path(excel_path).resolve()
    pdf_abs = Path(pdf_path).resolve()
    outdir = str(pdf_abs.parent)

    cmd = [
        office_bin,
        "--headless",
        "--convert-to",
        "pdf",
        "--outdir",
        outdir,
        str(excel_abs),
    ]
    proc = subprocess.run(cmd, capture_output=True, text=True)
    if proc.returncode != 0:
        raise RuntimeError(
            "LibreOffice PDF conversion failed.\n"
            f"Command: {' '.join(cmd)}\n"
            f"stdout: {proc.stdout}\n"
            f"stderr: {proc.stderr}"
        )

    generated = Path(outdir) / f"{excel_abs.stem}.pdf"
    if not generated.exists():
        raise RuntimeError("LibreOffice finished without output PDF.")
    if generated.resolve() != pdf_abs:
        generated.replace(pdf_abs)
    return "libreoffice"


def main() -> None:
    load_env_file(".env")

    base_url = getenv_required("REDMINE_BASE_URL")
    api_key = getenv_required("REDMINE_API_KEY")
    user_id = getenv_required("REDMINE_USER_ID")

    today = date.today()
    date_from = ""
    date_to = ""
    pdf_out = os.getenv("PDF_OUT", "filled.pdf")
    picked = choose_month_windows_tk(today, pdf_out)
    if picked is not None:
        date_from, date_to, pdf_out = picked
    else:
        date_from, date_to = resolve_date_range(today)

    excel_in = str(resolve_path(os.getenv("EXCEL_IN", "unfilled.xlsx")))
    excel_out = str(resolve_path(os.getenv("EXCEL_OUT", "filled.xlsx")))
    sheet_name = os.getenv("SHEET_NAME", "Munka1")
    arrival_time = os.getenv("ARRIVAL_TIME", "09:00")
    print_area = os.getenv("PRINT_AREA", "A1:F47")
    pdf_out = choose_pdf_save_path_windows(pdf_out)
    pdf_out = str(resolve_path(pdf_out))

    entries = get_time_entries(base_url, api_key, user_id, date_from, date_to)
    days = aggregate_hours_by_day(entries)
    effective_print_area = fill_excel(excel_in, excel_out, sheet_name, days, arrival_time, print_area)
    pdf_engine = export_pdf(excel_out, pdf_out, sheet_name, effective_print_area)

    print(f"Done. Created: {excel_out}")
    print(f"Done. Created: {pdf_out}")
    print(f"PDF engine: {pdf_engine}")
    print(f"Date range: {date_from} -> {date_to}")
    print(f"Print area: {effective_print_area}")
    print(f"Fetched entries: {len(entries)}")
    print(f"Filled workdays: {len(days)}")


if __name__ == "__main__":
    main()




